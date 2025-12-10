"""
MATRIZ DE DECISIÓN TÁCTICA - GENERADOR EXCEL
Genera hoja interactiva con checkboxes para gestión de crisis:
- Checkboxes para síntomas (Pulso Alto, Pensamiento Rumiante, etc.)
- Lógica condicional que desbloquea soluciones
- Enlaces a audios/videos específicos

Requisitos: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def crear_matriz_decision_tactica(output_path='../Data/matriz_decision_tactica.xlsx'):
    """
    Genera Excel con matriz de decisión táctica interactiva
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz de Decisión"
    
    # Estilos
    title_font = Font(size=16, bold=True, color="1F4788")
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    danger_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    success_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    
    # Título
    ws['A1'] = "🎯 MATRIZ DE DECISIÓN TÁCTICA - GESTIÓN DE CRISIS"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    # Subtítulo
    ws['A2'] = "Marca los síntomas que experimentas y el sistema te mostrará la solución"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # ==================== SECCIÓN 1: SÍNTOMAS FÍSICOS ====================
    ws['A4'] = "1️⃣ SÍNTOMAS FÍSICOS"
    ws['A4'].font = Font(size=12, bold=True)
    ws['A4'].fill = header_fill
    ws['A4'].font = header_font
    ws.merge_cells('A4:B4')
    
    sintomas_fisicos = [
        ("Pulso Alto (>100 bpm)", "B5"),
        ("Temblor en las manos", "B6"),
        ("Sudoración excesiva", "B7"),
        ("Tensión muscular", "B8"),
        ("Respiración acelerada", "B9")
    ]
    
    row = 5
    for sintoma, cell in sintomas_fisicos:
        ws[f'A{row}'] = sintoma
        ws[cell] = False  # Checkbox (se implementa con validación de datos)
        
        # Validación de datos para checkbox
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(cell)
        
        row += 1
    
    # ==================== SECCIÓN 2: SÍNTOMAS MENTALES ====================
    ws['A11'] = "2️⃣ SÍNTOMAS MENTALES"
    ws['A11'].font = Font(size=12, bold=True)
    ws['A11'].fill = header_fill
    ws['A11'].font = header_font
    ws.merge_cells('A11:B11')
    
    sintomas_mentales = [
        ("Pensamiento Rumiante", "B12"),
        ("Miedo al Fallo", "B13"),
        ("Parálisis por Análisis", "B14"),
        ("Obsesión por el 10", "B15"),
        ("Autocrítica excesiva", "B16")
    ]
    
    row = 12
    for sintoma, cell in sintomas_mentales:
        ws[f'A{row}'] = sintoma
        ws[cell] = False
        
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(cell)
        
        row += 1
    
    # ==================== SECCIÓN 3: DIAGNÓSTICO AUTOMÁTICO ====================
    ws['D4'] = "🔍 DIAGNÓSTICO"
    ws['D4'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['D4'].fill = danger_fill
    ws.merge_cells('D4:E4')
    
    # Lógica condicional: Pulso Alto + Pensamiento Rumiante
    ws['D6'] = "Crisis Nivel 1:"
    ws['E6'] = '=IF(AND(B5=TRUE,B12=TRUE),"⚠️ ACTIVADO","---")'
    ws['E6'].font = Font(bold=True)
    
    ws['D7'] = "Crisis Nivel 2:"
    ws['E7'] = '=IF(AND(B5=TRUE,B13=TRUE,B14=TRUE),"🚨 CRÍTICO","---")'
    ws['E7'].font = Font(bold=True)
    
    # ==================== SECCIÓN 4: SOLUCIONES ====================
    ws['D9'] = "💊 SOLUCIONES DESBLOQUEADAS"
    ws['D9'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['D9'].fill = success_fill
    ws.merge_cells('D9:E9')
    
    # Solución 1: Pulso Alto + Pensamiento Rumiante
    ws['D11'] = "Si Crisis Nivel 1:"
    ws['E11'] = '=IF(AND(B5=TRUE,B12=TRUE),"🎧 Protocolo de Veto","Bloqueado")'
    ws['E11'].font = Font(bold=True, color="0000FF", underline="single")
    
    # Solución 2: Miedo al Fallo
    ws['D12'] = "Si Miedo al Fallo:"
    ws['E12'] = '=IF(B13=TRUE,"🎧 Reset Mental Integrado","Bloqueado")'
    ws['E12'].font = Font(bold=True, color="0000FF", underline="single")
    
    # Solución 3: Parálisis por Análisis
    ws['D13'] = "Si Parálisis Análisis:"
    ws['E13'] = '=IF(B14=TRUE,"🎯 Drill Velocidad Absurda","Bloqueado")'
    ws['E13'].font = Font(bold=True, color="0000FF", underline="single")
    
    # Solución 4: Respiración (siempre disponible si pulso alto)
    ws['D14'] = "Si Pulso Alto:"
    ws['E14'] = '=IF(B5=TRUE,"🫁 Respiración 4-7-8","Bloqueado")'
    ws['E14'].font = Font(bold=True, color="0000FF", underline="single")
    
    # ==================== SECCIÓN 5: ENLACES A RECURSOS ====================
    ws['A18'] = "📚 BIBLIOTECA DE RECURSOS"
    ws['A18'].font = Font(size=12, bold=True)
    ws['A18'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    ws['A18'].font = Font(color="FFFFFF", bold=True)
    ws.merge_cells('A18:E18')
    
    recursos = [
        ("🎧 Protocolo de Veto", "../Media/audios/protocolo_veto.mp3"),
        ("🎧 Reset Mental Integrado", "../Media/audios/reset_mental_integrado.mp3"),
        ("🎧 Respiración 4-7-8", "../Media/audios/respiracion_4_7_8.mp3"),
        ("🎯 Drill Velocidad Absurda", "../Media/videos/drill_velocidad_absurda.mp4"),
        ("🎯 Drill Dedo Muerto", "../Media/videos/drill_dedo_muerto.mp4")
    ]
    
    row = 19
    for nombre, ruta in recursos:
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = ruta
        ws[f'B{row}'].font = Font(color="0000FF", underline="single")
        row += 1
    
    # ==================== INSTRUCCIONES ====================
    ws_instrucciones = wb.create_sheet("Instrucciones")
    
    instrucciones = [
        "📋 CÓMO USAR LA MATRIZ DE DECISIÓN TÁCTICA",
        "",
        "OBJETIVO:",
        "Esta herramienta te ayuda a identificar tu estado mental/físico",
        "y te proporciona la solución específica automáticamente.",
        "",
        "PASO 1: IDENTIFICAR SÍNTOMAS",
        "- Marca TRUE en los síntomas que experimentas",
        "- Puedes marcar varios a la vez",
        "",
        "PASO 2: VER DIAGNÓSTICO",
        "- El sistema detecta automáticamente el nivel de crisis",
        "- Crisis Nivel 1: Pulso Alto + Pensamiento Rumiante",
        "- Crisis Nivel 2: Múltiples síntomas combinados",
        "",
        "PASO 3: ACCEDER A SOLUCIONES",
        "- Las soluciones se DESBLOQUEAN automáticamente",
        "- Solo verás las soluciones relevantes para tus síntomas",
        "- Haz clic en el enlace para acceder al recurso",
        "",
        "EJEMPLO DE USO:",
        "1. Estás en competición",
        "2. Notas que tu pulso está alto (>100)",
        "3. No puedes dejar de pensar en el resultado",
        "4. Marcas: Pulso Alto = TRUE",
        "5. Marcas: Pensamiento Rumiante = TRUE",
        "6. El sistema muestra: '🎧 Protocolo de Veto'",
        "7. Haces clic y escuchas el audio",
        "",
        "PROTOCOLOS ESPECÍFICOS:",
        "",
        "🎧 PROTOCOLO DE VETO:",
        "- Cuando: Pulso Alto + Pensamiento Rumiante",
        "- Acción: Salir del puesto, resetear, volver",
        "- Duración: 3-5 minutos",
        "",
        "🎧 RESET MENTAL INTEGRADO:",
        "- Cuando: Miedo al Fallo o Bloqueo",
        "- Acción: Audio de reset completo",
        "- Duración: 5 minutos",
        "",
        "🫁 RESPIRACIÓN 4-7-8:",
        "- Cuando: Pulso Alto (>100 bpm)",
        "- Acción: Inhala 4s, Retén 7s, Exhala 8s",
        "- Repetir: 3-5 ciclos",
        "",
        "🎯 DRILL VELOCIDAD ABSURDA:",
        "- Cuando: Parálisis por Análisis",
        "- Acción: Rutina al doble de velocidad",
        "- Objetivo: Romper el bucle mental",
        "",
        "⚠️ IMPORTANTE:",
        "- Esta matriz NO sustituye al entrenador",
        "- Es una herramienta de EMERGENCIA",
        "- Si los síntomas persisten, consulta con tu coach",
        "",
        "💡 FUENTES:",
        "- Guía de Emergencia (3 Errores Mentales Capitales)",
        "- Metodología Buceta (CPRD)",
        "- Neurociencia Bachrach"
    ]
    
    for i, line in enumerate(instrucciones, start=1):
        ws_instrucciones[f'A{i}'] = line
        if line.startswith(('📋', '🎧', '🫁', '🎯', '⚠️', '💡')):
            ws_instrucciones[f'A{i}'].font = Font(bold=True, size=11)
        elif line.startswith(('PASO', 'OBJETIVO', 'EJEMPLO', 'PROTOCOLOS')):
            ws_instrucciones[f'A{i}'].font = Font(bold=True, color="1F4788")
    
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # Ajustar anchos en hoja principal
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    
    # Guardar
    wb.save(output_path)
    print(f"✅ Matriz de Decisión Táctica creada: {output_path}")
    return output_path


if __name__ == "__main__":
    crear_matriz_decision_tactica()
